"""
Excel Generator Engine for VAVE AI
Generates properly formatted Excel files with appropriate styling and EMBEDDED IMAGES.
"""

import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import logging

logger = logging.getLogger(__name__)

# Define the project base directory to resolve relative image paths
BASE_DIR = Path(__file__).parent.resolve()

def _resolve_image_path(image_url_or_path):
    """
    Converts a URL or relative path from the table data into an absolute system path.
    Example: "/static/generated/foo.jpg" -> "C:/Project/static/generated/foo.jpg"
    """
    if not image_url_or_path or image_url_or_path == "N/A":
        return None
    if str(image_url_or_path).strip().upper() in ("NAN", "NONE", ""):
        return None

    # Clean the path string
    clean_path = str(image_url_or_path).strip()
    
    # Remove URL parameters if any (e.g., ?v=1)
    if "?" in clean_path:
        clean_path = clean_path.split("?")[0]

    # Remove leading slash if present to make it joinable
    if clean_path.startswith("/") or clean_path.startswith("\\"):
        clean_path = clean_path[1:]

    # Construct absolute path assuming BASE_DIR is the project root
    abs_path = BASE_DIR / clean_path

    if abs_path.exists() and abs_path.is_file():
        return str(abs_path)
    
    return None

def generate_excel_from_table(table_data, output_folder=None):
    """
    Generates a properly formatted Excel file from table data with embedded images.
    
    Args:
        table_data: List of dictionaries containing the data to export
        output_folder: Path to folder where Excel file should be saved (optional)
    
    Returns:
        str: Filename of the generated Excel file, or None if generation failed
    """
    if not table_data:
        logger.warning("No table data provided for Excel generation")
        return None
    
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Cost Reduction Ideas"
        
        # --- CONFIGURATION ---
        # 1. Defined Column Sequence (aligned with Analysis Results table in app). CAPEX = detailed/fine estimation for tooling, inventory, and all such setup cost.
        DESIRED_COLUMNS = [
            "Idea Id",
            "Current Scenario Image",
            "Competitor Image",
            "Proposal Scenario Image",
            "Cost Reduction Idea",
            "Way Forward",
            "CAPEX",
            "Saving Value (INR)",
            "Weight Saving (Kg)",
            "Status",
            "Feasibility Score",
            "Cost Saving Score",
            "Weight Reduction Score",
            "Homologation Feasibility Score",
            "Homologation Theory",
            "Responsibility",
            "Origin",
            "Date",
            "Dept",
            "Validation Notes",
        ]

        # 2. Identify which columns contain images
        IMAGE_COLUMNS = [
            "Current Scenario Image",
            "Competitor Image",
            "Proposal Scenario Image",
        ]

        # --- STYLING ---
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        data_font = Font(size=10, name="Calibri")
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        border_style = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # --- WRITE HEADERS ---
        for col_idx, header in enumerate(DESIRED_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style
            
            # Set specific widths for image columns and long-text columns
            col_letter = get_column_letter(col_idx)
            if header in IMAGE_COLUMNS:
                ws.column_dimensions[col_letter].width = 20  # Width for images
            elif header in ("Cost Reduction Idea", "Way Forward", "Homologation Theory"):
                ws.column_dimensions[col_letter].width = 40  # Wider for long text
            elif header == "Validation Notes":
                ws.column_dimensions[col_letter].width = 35
            else:
                ws.column_dimensions[col_letter].width = 15

        # Set Header Row Height
        ws.row_dimensions[1].height = 30

        # --- WRITE DATA ROWS ---
        for row_idx, row_data in enumerate(table_data, start=2):
            # Set row height to accommodate images (approx 80 points)
            ws.row_dimensions[row_idx].height = 80
            
            for col_idx, header in enumerate(DESIRED_COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border_style
                
                # Retrieve value (handle missing keys gracefully)
                # Map frontend keys (Title Case), app.formatted keys, or agent/db keys (snake_case) to Desired Columns
                val = row_data.get(header)
                if val is None:
                    # Fallback strategies for mismatched keys
                    if header == "Idea Id":
                        val = row_data.get("idea_id") or row_data.get("Idea ID")
                    elif header == "Cost Reduction Idea":
                        val = row_data.get("cost_reduction_idea")
                    elif header == "CAPEX":
                        val = row_data.get("CAPEX") or row_data.get("capex") or row_data.get("Investment (Cr)") or row_data.get("investment_cr")
                    elif header == "Saving Value (INR)":
                        val = row_data.get("saving_value_inr")
                    elif header == "Weight Saving (Kg)":
                        val = row_data.get("weight_saving")
                    elif header == "Current Scenario Image":
                        val = row_data.get("current_scenario_image")
                    elif header == "Competitor Image":
                        val = row_data.get("competitor_image")
                    elif header == "Proposal Scenario Image":
                        val = row_data.get("proposal_scenario_image")
                    elif header == "Way Forward":
                        val = row_data.get("way_forward")
                    elif header == "Responsibility":
                        val = row_data.get("resp") or row_data.get("Responsibility")
                    elif header == "Date":
                        val = row_data.get("target_date")
                    elif header == "Dept":
                        val = row_data.get("dept")
                    elif header == "Status":
                        val = row_data.get("status") or row_data.get("validation_status")
                    elif header == "Origin":
                        val = row_data.get("origin") or "Existing Database"
                    elif header == "Feasibility Score":
                        val = row_data.get("feasibility_score")
                    elif header == "Cost Saving Score":
                        val = row_data.get("cost_saving_score")
                    elif header == "Weight Reduction Score":
                        val = row_data.get("weight_reduction_score")
                    elif header == "Homologation Feasibility Score":
                        val = row_data.get("homologation_feasibility_score")
                    elif header == "Homologation Theory":
                        val = row_data.get("homologation_theory")
                    elif header == "Validation Notes":
                        val = row_data.get("validation_notes")

                if val is None:
                    val = ""
                # Sanitize NaN / numpy.nan / string "nan" for text cells (images handled separately)
                if header not in IMAGE_COLUMNS and val != "":
                    try:
                        if isinstance(val, float) and (val != val):  # NaN
                            val = ""
                        elif str(val).strip().lower() in ("nan", "none"):
                            val = ""
                    except Exception:
                        pass

                # --- IMAGE EMBEDDING LOGIC ---
                if header in IMAGE_COLUMNS:
                    # Resolve absolute path
                    img_path = _resolve_image_path(val)
                    
                    if img_path:
                        try:
                            img = XLImage(img_path)
                            # Resize image to fit cell (Keep aspect ratio if possible, or force dims)
                            img.width = 100
                            img.height = 100
                            
                            # Anchor image to the cell
                            img.anchor = cell.coordinate
                            ws.add_image(img)
                        except Exception as img_err:
                            logger.error(f"Failed to embed image {img_path}: {img_err}")
                            cell.value = "Image Error"
                            cell.alignment = data_alignment
                            cell.font = data_font
                    else:
                        cell.value = "No Image"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = Font(size=8, color="888888")

                # --- TEXT LOGIC ---
                else:
                    cell.value = str(val)
                    cell.alignment = data_alignment
                    cell.font = data_font

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cost_reduction_ideas_{timestamp}.xlsx"
        
        # Determine output path
        if output_folder:
            output_path = Path(output_folder) / filename
        else:
            output_path = Path(filename)
        
        # Ensure directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel file generated with images: {output_path}")
        
        return filename
        
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generate_excel_from_table_in_memory(table_data, upload_folder=None):
    """
    Wrapper for on-demand generation.
    """
    from pathlib import Path
    import tempfile
    
    try:
        # Use provided upload folder or create temp directory
        if upload_folder:
            output_folder = Path(upload_folder)
        else:
            output_folder = Path(tempfile.gettempdir()) / "vave_exports"
        
        output_folder.mkdir(parents=True, exist_ok=True)
        
        filename = generate_excel_from_table(table_data, output_folder=output_folder)
        
        if filename:
            file_path = output_folder / filename
            return filename, str(file_path)
        else:
            return None, None
            
    except Exception as e:
        logger.error(f"Error in generate_excel_from_table_in_memory: {str(e)}")
        return None, None