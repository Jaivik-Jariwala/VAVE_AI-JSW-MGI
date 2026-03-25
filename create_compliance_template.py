import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_engineering_template():
    # Define the headers based on VAVE / Engineering requirements
    headers = [
        # --- 1. IDENTIFICATION ---
        "System Group",         # e.g. Chassis, Powertrain, Body
        "Sub-System",           # e.g. Braking, Suspension
        "Part Name",            # e.g. Brake Rotor (Front)
        "Part Number",          # OEM Part Number
        "Vehicle Model",        # e.g. Hector Plus
        
        # --- 2. MG (CURRENT) SPECIFICATIONS ---
        "MG Material Grade",    # e.g. G3000 Grey Cast Iron
        "MG Manufacturing Process", # e.g. Sand Casting + Machining
        "MG Surface Treatment", # e.g. Geomet Coating, Zinc Plating
        "MG Weight (kg)",       # e.g. 5.2
        "MG Cost (INR)",        # e.g. 1200
        "MG Supplier Name",     # Optional
        "MG Sourcing",          # Local / Import
        
        # --- 3. COMPETITOR (BENCHMARK) SPECIFICATIONS ---
        "Competitor Model",     # e.g. Kia Seltos
        "Competitor Material",  # e.g. G3500 High Carbon
        "Competitor Process",   # e.g. Gravity Die Casting
        "Competitor Weight (kg)",
        "Competitor Cost (INR)",
        "Competitor Feature Diff", # e.g. "Has cooling fins", "Vented"
        
        # --- 4. ENGINEERING CRITERIA (THE "BRAIN") ---
        "Safety Critical Item?",       # YES / NO
        "Regulatory Standard (AIS/FMVSS)", # e.g. IS: 11852
        "Homologation Impact?",        # YES / NO
        "NVH Critical?",               # YES / NO (Noise Vibration Harshness)
        "Durability Target (cycles)",  # e.g. 300,000 cycles
        
        # --- 5. MATERIAL & DESIGN CONSTRAINTS ---
        "Max Operating Temp (C)",      # e.g. 800
        "Min Yield Strength (MPa)",    # e.g. 250
        "Corrosion Requirement (Hours)", # e.g. 720h SST
        "Dimensional Tolerance (mm)",  # e.g. +/- 0.05
        "Design Non-Negotiables",      # e.g. "Mounting PCD must remain 114.3"
        "Material Constraints",        # e.g. "No Aluminum allowed due to heat"
        
        # --- 6. VAVE / FUTURE OPPORTUNITY ---
        "Approved Alt Materials",      # e.g. "CFC, High Carbon Steel"
        "Potential Weight Save (Est)",
        "Potential Cost Save (Est)"
    ]
    
    # Create sample row for guidance
    sample_data = [
        "Chassis", "Braking", "Front Brake Disc", "2311-BA-A", "MG Hector",
        "Grey Cast Iron G3000", "Sand Casting", "Geomet 360", 5.2, 1450, "Brembo", "Local",
        "Kia Seltos", "High Carbon Steel", "Casting", 4.8, 1300, "Larger Vents",
        "YES", "IS: 11852 / FMVSS 105", "YES", "YES", "500,000",
        "700", "220", "480", "+/- 0.02", "Hat height fixed", "No Aluminum allowed due to heat",
        "High Carbon Alloy", "0.4 kg", "150 INR"
    ]
    
    # Generate Workbook
    file_path = "compliance_data_template.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed VAVE Input"
    
    # Write Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        
        # Style: Blue Header, White Text, Bold
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Smart column width
        if "Constraint" in header or "Non-Negotiable" in header:
            width = 35
        elif "Process" in header or "Material" in header:
            width = 25
        else:
            width = 18
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Write Sample Data
    for col_num, val in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = val
        cell.font = Font(italic=True, color="555555") # Italic to show it's an example
        cell.alignment = Alignment(wrap_text=True)

    # Freeze Header
    ws.freeze_panes = "A2"

    wb.save(file_path)
    print(f"Detailed Template created: {os.path.abspath(file_path)}")

if __name__ == "__main__":
    create_engineering_template()
