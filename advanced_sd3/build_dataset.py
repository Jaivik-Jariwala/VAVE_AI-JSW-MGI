import openpyxl
from openpyxl_image_loader import SheetImageLoader
import pandas as pd
import os
import json

excel_path = r'd:\Internship Workspaces\JSW Morris Garages India Workspace\Deployment Project\VAVE_AI-JSW-MGI (5)\VAVE_AI-JSW-MGI - Copy\VAVE_AI-JSW-MGI - Copy\20260219_AIML_Data_Hector Plus 1.5 Gas CVT_Idea_Consolidated_W Image.xlsx_v1.xlsx'
output_dir = 'extracted_dataset'
os.makedirs(output_dir, exist_ok=True)
os.makedirs(os.path.join(output_dir, 'images'), exist_ok=True)

# Load workbook and sheet
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active # fallback to active
for s in wb.sheetnames:
    if 'Hector' in s or 'Idea' in s or 'Sheet1' in s:
        sheet = wb[s]
        break

try:
    image_loader = SheetImageLoader(sheet)
except Exception as e:
    print(f"Error initializing SheetImageLoader: {e}")
    image_loader = None

# Using pandas to get headers and text values easily
df = pd.read_excel(excel_path)
columns = df.columns.tolist()

try:
    img_col_idx = columns.index('Base Vehicle Image')
    base_img_col_letter = openpyxl.utils.get_column_letter(img_col_idx + 1)
except ValueError:
    base_img_col_letter = None

try:
    prompt_col_idx = columns.index('Cost Reduction Idea Proposal')
except ValueError:
    prompt_col_idx = None

records = []

for row_idx, row in df.iterrows():
    if pd.isna(row.get('Cost Reduction Idea Proposal')):
        continue
    
    prompt = str(row['Cost Reduction Idea Proposal'])
    idea_id = str(row.get('Idea Id', f'Idea_{row_idx}'))
    
    img_filename = None
    if image_loader and base_img_col_letter:
        cell_coord = f"{base_img_col_letter}{row_idx + 2}" # +2 because pandas is 0-indexed and excel skips header
        try:
            if image_loader.image_in(cell_coord):
                img = image_loader.get(cell_coord)
                img = img.convert('RGB')
                img_filename = f"{idea_id}.jpg"
                img.save(os.path.join(output_dir, 'images', img_filename))
        except Exception as e:
            print(f"Failed to extract image for {cell_coord}: {e}")

    records.append({
        'idea_id': idea_id,
        'image': img_filename,
        'prompt': prompt,
        'row_idx': row_idx
    })

with open(os.path.join(output_dir, 'metadata.jsonl'), 'w') as f:
    for rec in records:
        f.write(json.dumps(rec) + '\n')

print(f"Extracted {len(records)} records. Valid images: {sum(1 for r in records if r['image'] is not None)}")
